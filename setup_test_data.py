"""
Prepare test data directories and sample files for pipeline testing.

Usage:
    python setup_test_data.py

This script:
- Creates landing_zone directories for each OEM
- Provides instructions for adding test files
- Validates directory structure
"""

import sys
from pathlib import Path


def setup_directories():
    """Create required directories."""
    project_root = Path(__file__).parent
    accy_v2_dir = project_root / "accy_v2"

    directories = [
        accy_v2_dir / "data" / "landing_zone" / "mitsubishi",
        accy_v2_dir / "data" / "landing_zone" / "mazda",
        project_root / "accy_v2" / "output" / "dq_reports" / "mitsubishi",
        project_root / "accy_v2" / "output" / "dq_reports" / "mazda",
        project_root / "accy_v2" / "output" / "pipeline_logs" / "mitsubishi",
        project_root / "accy_v2" / "output" / "pipeline_logs" / "mazda",
        project_root / "accy_v2" / "output" / "ready_to_upload" / "mitsubishi",
        project_root / "accy_v2" / "output" / "ready_to_upload" / "mazda",
    ]

    print("Creating directories...")
    for directory in directories:
        directory.mkdir(parents=True, exist_ok=True)
        print(f"  [OK] {directory.relative_to(project_root)}")

    print("\n[OK] All directories ready\n")


def create_sample_mitsubishi():
    """Create sample Mitsubishi Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[WARNING] openpyxl not installed. Install with: pip install openpyxl")
        return False

    project_root = Path(__file__).parent
    accy_v2_dir = project_root / "accy_v2"
    output_file = accy_v2_dir / "data" / "landing_zone" / "mitsubishi" / "2026_Outlander_ES_EN.xlsx"

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026_Outlander_ES_EN"

    # Row 1: Model name
    ws["A1"] = "Outlander ES"
    ws["A1"].font = Font(bold=True, size=14)

    # Row 2: Headers
    headers = ["Part Number", "Description", "Description FR", "MSRP", "DNP", "Install Time", "Labour Rate", "Remarks", "Photo", "ES", "SEL", "Limited"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Row 3+: Sample data
    sample_data = [
        ["ACC-001", "Floor Mats", "Tapis de sol", 89.99, 45.00, 0.5, 25.00, "Genuine OEM", "", "X", "", ""],
        ["ACC-002", "Roof Rack", "Galerie de toit", 199.99, 100.00, 2.0, 50.00, "3-piece set", "", "X", "X", "X"],
        ["ACC-003", "Splash Guards", "Garde-boue", 79.99, 40.00, 0.75, 30.00, "All trims", "", "X", "X", "X"],
        ["ACC-004", "Hitch Cover", "Couvre-attelage", 49.99, 25.00, 0.25, 20.00, "Stainless steel", "", "", "X", "X"],
        ["ACC-005", "Mudflaps", "Bavettes", 69.99, 35.00, 1.0, 35.00, "Premium", "", "X", "X", ""],
    ]

    for row_idx, row_data in enumerate(sample_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if col_idx <= 7 or col_idx == 8:  # Data columns
                if col_idx in [4, 5, 6, 7]:  # Numeric columns
                    cell.number_format = "0.00"
            elif col_idx > 9:  # Trim columns
                cell.alignment = Alignment(horizontal="center")

    # Set column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 12
    for col in ["J", "K", "L"]:
        ws.column_dimensions[col].width = 8

    wb.save(output_file)
    print(f"[OK] Created sample file: {output_file.relative_to(project_root)}")
    return True


def create_sample_mazda():
    """Create sample Mazda CSV file."""
    try:
        import csv
    except ImportError:
        print("[WARNING] CSV module not available")
        return False

    project_root = Path(__file__).parent
    accy_v2_dir = project_root / "accy_v2"
    output_file = accy_v2_dir / "data" / "landing_zone" / "mazda" / "mazda_sample.csv"

    # Sample data
    headers = [
        "ModelYear",
        "CarLineCode",
        "TrimLevel",
        "PartNumber",
        "AccessoryName",
        "AccessoryNameFR",
        "MSRP",
        "StandardInstallTime",
        "RestrictionEN",
        "RestrictionFR",
    ]

    rows = [
        [2026, "CX-90", "ES", "P001", "Floor Mats", "Tapis de sol", 89.99, 0.5, "", ""],
        [2026, "CX-90", "ES", "P002", "Roof Rack", "Galerie de toit", 199.99, 2.0, "", ""],
        [2026, "CX-90", "SEL", "P003", "Splash Guards", "Garde-boue", 79.99, 0.75, "", ""],
        [2026, "CX-90", "PREFERRED", "P004", "Hitch Cover", "Couvre-attelage", 49.99, 0.25, "", ""],
        [2026, "CX-90", "PREFERRED", "P005", "Mudflaps", "Bavettes", 69.99, 1.0, "", ""],
    ]

    try:
        with open(output_file, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)
        print(f"[OK] Created sample file: {output_file.relative_to(project_root)}")
        return True
    except Exception as e:
        print(f"[ERROR] Failed to create Mazda sample: {e}")
        return False


def main():
    """Main setup."""
    print("\n" + "="*80)
    print("TEST DATA SETUP")
    print("="*80 + "\n")

    # Create directories
    setup_directories()

    # Create sample files
    print("Creating sample data files...")
    try:
        create_sample_mitsubishi()
    except Exception as e:
        print(f"[WARNING] Failed to create Mitsubishi sample: {e}")

    try:
        create_sample_mazda()
    except Exception as e:
        print(f"[WARNING] Failed to create Mazda sample: {e}")

    # Print instructions
    print("\n" + "="*80)
    print("NEXT STEPS")
    print("="*80 + "\n")

    print("1. Populate the vehicle model database (if not already done):")
    print("   python -c \"\"\"")
    print("   import sys; from pathlib import Path")
    print("   sys.path.insert(0, str(Path('model_lookup')))")
    print("   from model_lookup.engine import load_env, create_engine_from_env")
    print("   from model_lookup.models.manufacture_module import batch_save_manufacturer_models")
    print("   load_env()")
    print("   engine = create_engine_from_env()")
    print("   batch_save_manufacturer_models(engine, ['Mitsubishi', 'Mazda'], 'model_lookup/db/db_vehicle_models.csv')")
    print("   \"\"\"")

    print("\n2. Test the pipeline:")
    print("   python test_model_lookup_pipeline.py mitsubishi")
    print("   python test_model_lookup_pipeline.py mazda")

    print("\n3. Review results:")
    print("   - DQ Report: accy_v2/output/dq_reports/{oem}/*.json")
    print("   - Pipeline Log: accy_v2/output/pipeline_logs/{oem}/*.log")
    print("   - Output Files: accy_v2/output/ready_to_upload/{oem}/*.xlsx")

    print("\n" + "="*80)
    print("[OK] Setup complete!\n")


if __name__ == "__main__":
    main()
