# ---
# jupyter:
#   jupytext:
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.19.1
#   kernelspec:
#     display_name: base
#     language: python
#     name: python3
# ---

# %%

import pandas as pd
import pathlib as pl
import os
from glob import glob


from __future__ import annotations

import re
import logging
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Optional, Dict, List, Tuple, Union

import numpy as np


from openpyxl import load_workbook
from openpyxl.styles import PatternFill
 

# %%

# %%
def dbg_print_df_columns(df: pd.DataFrame, title: str, model_key: str, section_key: str, max_cols: int = 80):
    print("\n" + "="*100)
    print(f"[DEBUG] {title}")
    print(f"Model: {model_key} | Section: {section_key}")
    print(f"Shape: {df.shape}")
    cols = list(df.columns)
    print(f"Columns ({len(cols)}):")
    for i, c in enumerate(cols[:max_cols]):
        print(f"  {i}: {repr(c)}")
    if len(cols) > max_cols:
        print(f"  ... {len(cols)-max_cols} more")
    print("="*100)


def dbg_print_section_header_rows(section_df: pd.DataFrame, title: str, model_key: str, section_key: str, rows: int = 3, cols: int = 25):
    print("\n" + "-"*100)
    print(f"[DEBUG] {title} (first {rows} rows x first {cols} cols)")
    print(f"Model: {model_key} | Section: {section_key} | Raw shape: {section_df.shape}")
    # show raw header rows used for header build
    preview = section_df.iloc[:rows, :min(cols, section_df.shape[1])].copy()
    print(preview)
    print("-"*100)


def dbg_find_comments_like_columns(df: pd.DataFrame, title: str, model_key: str, section_key: str):
    print("\n" + "~"*100)
    print(f"[DEBUG] {title} - Comments-like column search")
    print(f"Model: {model_key} | Section: {section_key}")
    candidates = []
    for c in df.columns:
        cn = str(c).strip().lower()
        cn_norm = cn.replace(":", "").replace("(s)", "s")
        if "comment" in cn_norm:
            candidates.append(c)
    print(f"Found {len(candidates)} candidates:")
    for c in candidates:
        print(f"  - {repr(c)}")
    print("~"*100)

# %%



# Definition
VALID_APPLICABILITY = {"•", "T", "E", "S"}

# =============================================================================
# Logging
# =============================================================================
def _build_logger(name: str = "accessory_pipeline") -> logging.Logger:
    logger = logging.getLogger(name)
    if not logger.handlers:
        handler = logging.StreamHandler()
        fmt = logging.Formatter("[%(levelname)s] %(message)s")
        handler.setFormatter(fmt)
        logger.addHandler(handler)
    logger.setLevel(logging.INFO)
    return logger


# =============================================================================
# Exceptions
# =============================================================================
class PipelineError(Exception):
    pass


class WorksheetNotFoundError(PipelineError):
    def __init__(self, model_key: str, worksheet_name: str, available_sheets: List[str]):
        super().__init__(
            f"Worksheet '{worksheet_name}' not found for model '{model_key}'. "
            f"Available sheets: {available_sheets}"
        )


class DataValidationError(PipelineError):
    def __init__(self, message: str, model_key: Optional[str] = None, section_key: Optional[str] = None, details: Any = None):
        prefix = []
        if model_key is not None:
            prefix.append(f"model='{model_key}'")
        if section_key is not None:
            prefix.append(f"section='{section_key}'")
        p = f"[{', '.join(prefix)}] " if prefix else ""
        super().__init__(p + message + (f" | details={details}" if details is not None else ""))


class SectionParseError(PipelineError):
    def __init__(self, section_key: str, reason: str, context: Any = None):
        super().__init__(f"Section parse error for '{section_key}': {reason}" + (f" | context={context}" if context is not None else ""))


# =============================================================================
# Utilities
# =============================================================================
PRICING_IDENTITY_COLS = [
    "description",
    "part number",
    "dealer net",
    "list price",
    "frt",
    "installed price",
]


METADATA_COLUMNS = {
    "model_name",
    "model_year",
    "publication_date",
    "make",
}



# Detect data row (safety to avoid deleting real data)
_PARTNO_RX = re.compile(r"[A-Z0-9]{2,}[-][A-Z0-9-]{2,}", flags=re.IGNORECASE)
_NUM_RX = re.compile(r"^\d+(\.\d+)?$")


def validate_required_columns_not_empty(
    df: pd.DataFrame,
    required_cols: list,
    model_key: str,
    section_key: str,
):
    if df.empty:
        return

    violations = []

    for col in required_cols:
        empty_mask = df[col].map(lambda v: clean_cell(v) == "")
        if empty_mask.any():
            rows = df.loc[empty_mask, PRICING_IDENTITY_COLS].to_dict(orient="records")
            violations.append({
                "column": col,
                "affected_rows": rows[:5],  # sample
            })

    if violations:
        raise DataValidationError(
            "Required pricing columns contain empty values.",
            model_key=model_key,
            section_key=section_key,
            details={"violations": violations},
        )

def validate_pricing_identity_uniqueness(
    df: pd.DataFrame,
    model_key: str,
    section_key: str,
    core_columns: list,
):
    """
    Ensures no duplicate rows exist with identical:
      - pricing identity
      - applicability signature
    """

    if df.empty:
        return

    dynamic_cols = [
        c for c in df.columns
        if c.lower() not in {c.lower() for c in core_columns}
    ]

    df = df.copy()

    # Pricing identity
    df["_pricing_key"] = df[PRICING_IDENTITY_COLS].astype(str).agg("|".join, axis=1)

    # Applicability signature
    df["_app_sig"] = df.apply(
        lambda r: tuple(
            
            (r.get(c, "")) or None
            for c in dynamic_cols
        ),
        axis=1
    )

    dupes = (
        df.groupby(["_pricing_key", "_app_sig"])
        .size()
        .reset_index(name="count")
        .query("count > 1")
    )

    if not dupes.empty:
        raise DataValidationError(
            "Duplicate rows with identical pricing and applicability detected.",
            model_key=model_key,
            section_key=section_key,
            details={
                "duplicates": dupes.to_dict(orient="records")
            }
        )
def _is_nullish(x: Any) -> bool:
    if x is None:
        return True
    if isinstance(x, float) and np.isnan(x):
        return True
    if pd.isna(x):
        return True
    if isinstance(x, str) and x.strip() == "":
        return True
    return False

def guard_dynamic_columns(
    *,
    before_df: pd.DataFrame,
    after_df: pd.DataFrame,
    core_columns: List[str],
    model_key: str,
    section_key: str,
    stage: str,
):
    """
    Guardrail to ensure dynamic (non-core) columns did not lose REAL data
    between pipeline stages.

    What this guard checks:
    - Only dynamic columns (not core columns)
    - Compares before vs after at the COLUMN level
    - Flags a column if it had real data before and has NONE after

    Definitions:
    - Core columns: provided explicitly (never checked here)
    - Real data:
        * Any non-empty value
        * Including applicability markers: •, T, E, S
        * Empty is meaningful, but total wipe is NOT allowed

    This guard DOES NOT:
    - Enforce column existence
    - Enforce uniqueness
    - Drop columns
    """

    if before_df.empty or after_df.empty:
        # Nothing to compare safely
        return

    core_lower = {c.lower() for c in core_columns}

    lost_columns = []

    for col in before_df.columns:
        # ✅ Skip core columns
        if col.lower() in core_lower:
            continue

        # Column must exist in after_df to be comparable
        if col not in after_df.columns:
            lost_columns.append({
                "column": col,
                "reason": "column_missing_after_stage",
            })
            continue

        # Count non-empty values BEFORE
        before_non_empty = before_df[col].map(
            lambda v: clean_cell(v) != ""
        ).sum()

        # Count non-empty values AFTER
        after_non_empty = after_df[col].map(
            lambda v: clean_cell(v) != ""
        ).sum()

        # ✅ If column had data before and none after → REAL DATA LOSS
        if before_non_empty > 0 and after_non_empty == 0:
            lost_columns.append({
                "column": col,
                "before_non_empty": int(before_non_empty),
                "after_non_empty": int(after_non_empty),
            })

    if lost_columns:
        raise DataValidationError(
            "Dynamic column data loss detected.",
            model_key=model_key,
            section_key=section_key,
            details={
                "stage": stage,
                "lost_columns": lost_columns,
            },
        )


def attach_model_metadata_to_df(
    df: pd.DataFrame,
    metadata: Dict[str, Any]
) -> pd.DataFrame:
    df = df.copy()
    df["model_name"] = metadata["model_name"]
    df["model_year"] = metadata["model_year"]
    df["make"] = metadata.get("make")
    return df


# %%


# Definition
VALID_APPLICABILITY = {"•", "T", "E", "S"}

# =============================================================================
# Logging
# =============================================================================


def _build_logger(name: str = "accessory_pipeline") -> logging.Logger:
    logger = logging.getLogger(name)
    if not logger.handlers:
        handler = logging.StreamHandler()
        fmt = logging.Formatter("[%(levelname)s] %(message)s")
        handler.setFormatter(fmt)
        logger.addHandler(handler)
    logger.setLevel(logging.INFO)
    return logger

LOGGER = _build_logger("accessory_pipeline")


# =============================================================================
# Exceptions (support details)
# =============================================================================
class PipelineError(Exception):
    pass

class WorksheetNotFoundError(PipelineError):
    def __init__(self, message: str, model_key: Optional[str] = None, details: Any = None):
        super().__init__(message + (f" | model={model_key}" if model_key else "") + (f" | details={details}" if details is not None else ""))

class DataValidationError(PipelineError):
    def __init__(self, message: str, model_key: Optional[str] = None, section_key: Optional[str] = None, details: Any = None):
        prefix = []
        if model_key is not None:
            prefix.append(f"model='{model_key}'")
        if section_key is not None:
            prefix.append(f"section='{section_key}'")
        p = f"[{', '.join(prefix)}] " if prefix else ""
        super().__init__(p + message + (f" | details={details}" if details is not None else ""))

class SectionParseError(PipelineError):
    def __init__(self, section_key: str, reason: str, context: Any = None):
        super().__init__(f"Section parse error for '{section_key}': {reason}" + (f" | context={context}" if context is not None else ""))


# =============================================================================
# Config & Result
# =============================================================================
@dataclass(frozen=True)
class PipelineConfig:
    # English worksheet name pattern (2 digits + optional underscore)
    english_pattern_template: str = r"^\d{2}_?MY_.*_APP_EN$"
    

    # Section keys
    key_file_sections: Tuple[str, ...] = (
        "1.0 Packages and Kits",
        "2.0 Wheels and Wheel Accessories",
        "3.0 Exterior",
        "4.0 Interior",
        "5.0 Electrical",
        "6.0 Paint Pens",
    )
    description_header_text: str = "Description"

    # Core columns (MUST exist; MUST NOT be dropped)
    required_columns: Tuple[str, ...] = (
        "description",
        "part number",
        "dealer net",
        "list price",
        "frt",
        "installed price",
        "comments",
    )

    # Starter column autodetect
    autodetect_scan_first_n_cols: int = 12

    # Dynamic header fill
    dynamic_col_prefix: str = "extra_col"

    # Required row cleanup pre-merge
    drop_description_only_rows_pre_merge: bool = True
    drop_pricing_marker_rows_pre_merge: bool = True
    pricing_marker_strings: Tuple[str, ...] = ("red pricing", "promotional pricing")  # in part number

    # Cleaning
    strip_strings_on_load: bool = True
    
    # Package behavior
    keep_package_children_rows: bool = False

    # Audit controls
    audit_enabled: bool = True
    audit_model_filter: Optional[str] = None
    audit_section_filter: Optional[str] = None

    # Strict
    strict_mode: bool = True


@dataclass
class ProcessedRawFileResult:
    model_key: str
    merged_df: pd.DataFrame
    english_sheet_name: Optional[str]
    section_dfs: Dict[str, pd.DataFrame] = field(default_factory=dict)


# =============================================================================
# Auditor
# =============================================================================
@dataclass
class ColumnSnapshot:
    stage: str
    shape: Tuple[int, int]
    columns: List[str]

class ColumnAuditor:
    def __init__(self, enabled: bool = True, model_filter: Optional[str] = None, section_filter: Optional[str] = None):
        self.enabled = enabled
        self.model_filter = model_filter
        self.section_filter = section_filter
        self.snapshots: Dict[Tuple[str, str], List[ColumnSnapshot]] = {}
        self.df_cache: Dict[Tuple[str, str, str], pd.DataFrame] = {}

    def _match(self, model: str, section: str) -> bool:
        if not self.enabled:
            return False
        if self.model_filter and model != self.model_filter:
            return False
        if self.section_filter and section != self.section_filter:
            return False
        return True

    def take(self, model: str, section: str, stage: str, df: pd.DataFrame, sample_removed: int = 3):
        if not self._match(model, section):
            return

        df2 = df.copy()
        cols = [str(c) for c in df2.columns]
        snap = ColumnSnapshot(stage=stage, shape=df2.shape, columns=cols)
        key = (model, section)
        self.snapshots.setdefault(key, []).append(snap)
        self.df_cache[(model, section, stage)] = df2

        print("\n" + "=" * 120)
        print(f"[AUDIT] model={model} section={section} stage={stage} shape={df2.shape} col_count={len(cols)}")
        print("Columns:", cols)

        snaps = self.snapshots[key]
        if len(snaps) >= 2:
            prev = snaps[-2]
            removed = [c for c in prev.columns if c not in cols]
            added = [c for c in cols if c not in prev.columns]
            if removed or added:
                print(f"\n[AUDIT DIFF] from '{prev.stage}' -> '{stage}'")
                print("  added:", added)
                print("  removed:", removed)

                prev_df = self.df_cache[(model, section, prev.stage)]
                for c in removed[:30]:
                    if c in prev_df.columns:
                        ser = prev_df[c]
                        non_empty = ser.map(lambda x: ("" if pd.isna(x) else str(x).strip()) != "").sum()
                        print(f"    removed_col={repr(c)} non_empty_count={int(non_empty)}")
                        if non_empty > 0:
                            vals = ser.dropna().astype(str)
                            vals = vals[vals.str.strip() != ""].head(sample_removed).tolist()
                            print(f"      sample_values: {vals}")

        print("=" * 120)

def make_auditor(config: PipelineConfig) -> ColumnAuditor:
    return ColumnAuditor(
        enabled=config.audit_enabled,
        model_filter=config.audit_model_filter,
        section_filter=config.audit_section_filter
    )


# =============================================================================
# Check column data
# =============================================================================

def has_applicability_signal(row: pd.Series, dynamic_cols: list) -> bool:
    for c in dynamic_cols:
        if clean_cell(row.get(c, "")) in {"•", "T", "E", "S"}:
            return True
    return False


# =============================================================================
# Common helpers
# =============================================================================
def clean_cell(x: Any) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)) or pd.isna(x):
        return ""
    s = str(x).replace("\r", " ").replace("\n", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s

def normalize_text(x: Any) -> str:
    s = clean_cell(x).lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("’", "'").replace("œ", "oe")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def strip_strings_in_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    obj_cols = out.select_dtypes(include=["object"]).columns
    if len(obj_cols) == 0:
        return out
    out[obj_cols] = out[obj_cols].apply(lambda col: col.map(lambda v: clean_cell(v) if isinstance(v, str) else v))
    return out

def _is_empty_like(x: Any) -> bool:
    # '.' and 'N/A' are NOT empty
    if x is None or (isinstance(x, float) and np.isnan(x)) or pd.isna(x):
        return True
    if isinstance(x, str):
        return x.strip() == ""
    return False

def col_all_empty(df: pd.DataFrame, col: str) -> bool:
    return df[col].map(_is_empty_like).all()

def col_all_zero_numeric(df: pd.DataFrame, col: str) -> bool:
    ser = pd.to_numeric(df[col], errors="coerce")
    return ser.notna().any() and (ser.fillna(0) == 0).all()

def count_distinct_applicability(col: pd.Series) -> int:
    """
    Counts distinct applicability states in a column.

    Valid states:
      - '•', 'T', 'E', 'S'
      - None (empty)

    Returns the number of distinct states found.
    """
    states = set()

    for v in col:
        v_clean = clean_cell(v)

        if v_clean == "":
            states.add(None)
        elif v_clean in VALID_APPLICABILITY:
            states.add(v_clean)

    return len(states)

def drop_only_truly_useless_columns(
    df: pd.DataFrame,
    protect: List[str],
) -> pd.DataFrame:
    """
    Drops non-core columns that do NOT have enough applicability signal.

    FINAL RULE (authoritative):
    - Core columns are NEVER dropped
    - Columns named Application* or 0.0* are dropped earlier (unchanged)
    - Dynamic columns are KEPT only if they contain
      >= 2 distinct values among:
        { '•', 'T', 'E', 'S', empty }

    This removes alignment artifacts (extra_col_*)
    while preserving real trim/applicability columns.
    """

    protect_lower = {p.lower() for p in protect}
    drop_cols = []

    for col_name in df.columns:
        # ✅ Never drop core columns
        if col_name.lower() in protect_lower:
            continue

        col = df[col_name]

        distinct_states = count_distinct_applicability(col)

        # ✅ KEEP only if there is real variability
        if distinct_states < 2:
            drop_cols.append(col_name)

    return df.drop(columns=drop_cols) if drop_cols else df

def _row_all_empty(series: pd.Series) -> bool:
    return all(clean_cell(v) == "" for v in series.tolist())

def _dedupe_columns(cols: List[str]) -> List[str]:
    seen = {}
    out = []
    for c in cols:
        if c not in seen:
            seen[c] = 1
            out.append(c)
        else:
            seen[c] += 1
            out.append(f"{c}_{seen[c]}")
    return out



def row_looks_like_data(row: pd.Series) -> bool:
    tokens = ["" if pd.isna(x) else str(x).strip() for x in row.tolist()]
    tokens = [t for t in tokens if t != ""]
    if not tokens:
        return False
    partno_hits = sum(1 for t in tokens if _PARTNO_RX.search(t))
    numeric_hits = sum(1 for t in tokens if _NUM_RX.match(t))
    return (partno_hits >= 1) or (numeric_hits >= 2)


# =============================================================================
# Worksheet & section locating
# =============================================================================
def find_english_worksheet_name(model_dict: Dict[str, pd.DataFrame], config: PipelineConfig, model_key: str) -> Optional[str]:
    model_escaped = re.escape(clean_cell(model_key))
    pattern = config.english_pattern_template.replace("{model}", model_escaped)
    rx = re.compile(pattern, flags=re.IGNORECASE)
    for name in model_dict.keys():
        if rx.match(clean_cell(name)):
            return name
    return None

def autodetect_starter_col(df: pd.DataFrame, config: PipelineConfig) -> Optional[int]:
    if df.empty:
        return None
    n_cols = min(config.autodetect_scan_first_n_cols, df.shape[1])
    targets = {normalize_text(k) for k in config.key_file_sections}

    best_col, best_hits = None, 0
    for col_idx in range(n_cols):
        hits = int(df.iloc[:, col_idx].map(normalize_text).isin(targets).sum())
        if hits > best_hits:
            best_hits = hits
            best_col = col_idx
    return best_col if best_hits > 0 else None

def extract_sections(model_key: str, english_df: pd.DataFrame, config: PipelineConfig, auditor: ColumnAuditor) -> Dict[str, pd.DataFrame]:
    starter_col = autodetect_starter_col(english_df, config)
    if starter_col is None:
        raise DataValidationError("Could not autodetect starter column for section keys.", model_key=model_key)

    starter_series = english_df.iloc[:, starter_col].map(normalize_text)
    section_to_idx: Dict[str, int] = {}

    for s in config.key_file_sections:
        hits = starter_series[starter_series == normalize_text(s)]
        if len(hits) == 0:
            LOGGER.warning(f"Section key not found: '{s}' | {{'model_key': '{model_key}'}}")
            continue
        section_to_idx[s] = int(hits.index[0])

    sections_sorted = sorted(section_to_idx.items(), key=lambda kv: kv[1])
    extracted: Dict[str, pd.DataFrame] = {}

    for i, (section_key, section_row_idx) in enumerate(sections_sorted):
        next_start = sections_sorted[i + 1][1] if i + 1 < len(sections_sorted) else english_df.shape[0]

        header_row = None
        for r in range(section_row_idx + 1, min(next_start, english_df.shape[0])):
            if normalize_text(english_df.iat[r, starter_col]) == normalize_text(config.description_header_text):
                header_row = r
                break

        if header_row is None:
            raise SectionParseError(section_key, "Could not find Description row.", {"model_key": model_key, "starter_col": starter_col})

        sec_df = english_df.iloc[header_row:next_start, :].copy()
        auditor.take(model_key, section_key, "extract_sections:raw_section_slice", sec_df)
        extracted[section_key] = sec_df

    return extracted


# =============================================================================
# Header logic per your new requirement
# =============================================================================
CORE_CANON = {
    "description": "description",
    "part number": "part number",
    "part no": "part number",
    "part no.": "part number",
    "part#": "part number",
    "part #": "part number",
    "dealer net": "dealer net",
    "list price": "list price",
    "frt": "frt",
    "installed price": "installed price",
    "comments": "comments",
    "comment": "comments",
    "comment(s)": "comments",
    "comments:": "comments",
}

APPLICABILITY_MARKERS = {"•", "T", "E", "S"}

def canonicalize_header_token(x: Any) -> str:
    s = normalize_text(x).replace(":", "")
    if s in CORE_CANON:
        return CORE_CANON[s]
    if s.startswith("comment"):
        return "comments"
    return s  # normalized token for matching

def find_header_row_with_all_core(section_df: pd.DataFrame, core_cols: List[str], scan_rows: int = 6) -> Optional[int]:
    core_set = set([c.lower() for c in core_cols])
    rmax = min(scan_rows, section_df.shape[0])
    for r in range(rmax):
        tokens = set(canonicalize_header_token(v) for v in section_df.iloc[r, :].tolist() if clean_cell(v) != "")
        if core_set.issubset(tokens):
            return r
    return None

def drop_columns_named_zero_or_application(df: pd.DataFrame) -> pd.DataFrame:
    """
    Drop columns where header equals '0.0' (or startswith '0.0') OR header startswith 'application'
    case-insensitive.
    """
    drop_cols = []
    for c in df.columns:
        name = normalize_text(c)
        if name == "0.0" or name.startswith("0.0"):
            drop_cols.append(c)
        if name == "application" or name.startswith("application"):
            drop_cols.append(c)
    drop_cols = list(dict.fromkeys(drop_cols))  # dedupe preserve order
    return df.drop(columns=drop_cols) if drop_cols else df

def fix_section_header_new_spec(model_key: str, section_key: str, section_df: pd.DataFrame,
                                config: PipelineConfig, auditor: ColumnAuditor) -> pd.DataFrame:
    """
    Implements your new requirement exactly.
    """
    df = section_df.copy()
    auditor.take(model_key, section_key, "fix_header:input_raw_section", df)

    # Drop leading empty column (label column) if truly empty
    if df.shape[1] > 0 and df.iloc[:10, 0].map(lambda x: ("" if pd.isna(x) else str(x).strip()) == "").all():
        df = df.drop(columns=[df.columns[0]]).reset_index(drop=True)
    auditor.take(model_key, section_key, "fix_header:after_drop_leading_empty_col", df)

    # Locate header row containing all core columns
    header_row_idx = find_header_row_with_all_core(df, list(config.required_columns), scan_rows=6)
    if header_row_idx is None:
        raise DataValidationError(
            "Could not locate header row containing all core columns.",
            model_key=model_key,
            section_key=section_key,
            details={"required_columns": list(config.required_columns)}
        )

    # Step 2a: Fill empty header cells from next row same column
    if header_row_idx + 1 < df.shape[0]:
        header_row = df.iloc[header_row_idx].copy()
        helper_row = df.iloc[header_row_idx + 1].copy()

        for c in range(df.shape[1]):
            hdr = clean_cell(header_row.iloc[c])
            helper = clean_cell(helper_row.iloc[c])
            if hdr == "" and helper != "" and helper not in ("0", "0.0"):
                header_row.iloc[c] = helper_row.iloc[c]

        df.iloc[header_row_idx] = header_row
        auditor.take(model_key, section_key, "fix_header:after_fill_empty_header_cells_from_next_row", df)

        # Step 2b: Delete second row (helper) ONLY if it doesn't look like data (safety guard)
        if not row_looks_like_data(helper_row):
            df = df.drop(index=[df.index[header_row_idx + 1]]).reset_index(drop=True)
            auditor.take(model_key, section_key, "fix_header:after_delete_helper_row", df)

    # Promote header_row_idx to header
    header_vals = [clean_cell(v) for v in df.iloc[header_row_idx].tolist()]
    # canonicalize only core headers; keep dynamic as-is (cleaned)
    new_cols = []
    for i, h in enumerate(header_vals):
        h_clean = clean_cell(h)
        h_norm = canonicalize_header_token(h_clean)
        # if this token is a core name, use canonical lowercase
        if h_norm in set(CORE_CANON.values()):
            new_cols.append(h_norm)
        else:
            # dynamic header: keep cleaned string; if still empty create extra_col_i
            new_cols.append(h_clean if h_clean != "" else f"{config.dynamic_col_prefix}_{i}")

    # ----------------------------------------
    # Finalize header names
    # ----------------------------------------
    new_cols = _dedupe_columns(new_cols)

    # ----------------------------------------
    # DROP header row + helper row (SAFE, ORDERED)
    # ----------------------------------------

    # Always drop the header row
    rows_to_drop = {header_row_idx}

    # Always drop the helper row if it exists
    helper_row_idx = header_row_idx + 1
    if helper_row_idx < len(df):
        rows_to_drop.add(helper_row_idx)

    # ✅ DEFINE keep_positions FIRST (UNCONDITIONAL)
    keep_positions = [i for i in range(len(df)) if i not in rows_to_drop]

    # ✅ NOW slice using keep_positions
    data_df = df.iloc[keep_positions].reset_index(drop=True)

    # ✅ NOW assign columns
    data_df.columns = new_cols

    # ----------------------------------------
    # Header guard (position-based, CORRECT)
    # ----------------------------------------
    guard_header_data_preservation(
        raw_df=df,
        data_df=data_df,
        core_columns=list(config.required_columns),
        model_key=model_key,
        section_key=section_key,
    )


    auditor.take(
        model_key,
        section_key,
        "fix_header:after_drop_header_and_helper_rows",
        data_df
    )

    data_df.columns = new_cols
    auditor.take(model_key, section_key, "fix_header:after_promote_header", data_df)

    # Drop columns named 0.0 or Application (and variants)
    data_df = drop_columns_named_zero_or_application(data_df)
    auditor.take(model_key, section_key, "fix_header:after_drop_0.0_and_application_cols", data_df)

    # Drop only truly useless columns BUT never drop core columns
    data_df = drop_only_truly_useless_columns(data_df, protect=list(config.required_columns))
    auditor.take(model_key, section_key, "fix_header:after_drop_only_truly_useless_columns(PROTECTED)", data_df)

    # Validate core columns exist
    cols_lower = [c.lower() for c in data_df.columns]
    missing = [c for c in config.required_columns if c.lower() not in cols_lower]
    if missing:
        raise DataValidationError(
            "Missing required/core columns after header processing.",
            model_key=model_key,
            section_key=section_key,
            details={"missing": missing, "columns": data_df.columns.tolist()}
        )

    return data_df



def column_has_applicability_data(col: pd.Series) -> bool:
    """
    Returns True if column contains any valid applicability marker:
    •, T, E, or S
    """
    return col.astype(str).str.strip().isin(APPLICABILITY_MARKERS).any()

def get_dynamic_columns_with_data(df: pd.DataFrame, core_columns: List[str]) -> Dict[str, int]:
    """
    Returns {column_name: non_empty_count} for dynamic columns that have data.
    """
    core_lower = {c.lower() for c in core_columns}
    out = {}

    for c in df.columns:
        if c.lower() in core_lower:
            continue
        non_empty = df[c].map(lambda v: clean_cell(v) != "").sum()
        if non_empty > 0:
            out[c] = int(non_empty)

    return out

def guard_header_data_preservation(
    *,
    raw_df: pd.DataFrame,
    data_df: pd.DataFrame,
    core_columns: List[str],
    model_key: str,
    section_key: str,
):
    """
    Guardrail: ensure header promotion did not wipe REAL data.

    Rules:
    - Compare by column POSITION (not name)
    - Ignore core columns (e.g. 'comments')
    - Ignore header rows in raw_df
    """
    core_lower = {c.lower() for c in core_columns}
    lost = []

    # Raw data rows = everything except first two rows (header + helper)
    raw_data_df = raw_df.iloc[2:].reset_index(drop=True)

    width = min(raw_data_df.shape[1], data_df.shape[1])

    for i in range(width):
        col_name = data_df.columns[i]

        # ✅ Skip core columns entirely
        if col_name.lower() in core_lower:
            continue

        before_non_empty = raw_data_df.iloc[:, i].map(
            lambda v: clean_cell(v) != ""
        ).sum()

        after_non_empty = data_df.iloc[:, i].map(
            lambda v: clean_cell(v) != ""
        ).sum()

        # ✅ Real loss: data rows existed before, none after
        if before_non_empty > 0 and after_non_empty == 0:
            lost.append({
                "column_position": i,
                "column_name": col_name,
                "before_non_empty": int(before_non_empty),
                "after_non_empty": int(after_non_empty),
            })

    if lost:
        raise DataValidationError(
            "Header promotion wiped real column data.",
            model_key=model_key,
            section_key=section_key,
            details={"lost_columns": lost},
        )


# =============================================================================
# Pre-merge cleanup rules (your requirements)
# =============================================================================
def drop_pricing_marker_rows(df: pd.DataFrame, markers: Tuple[str, ...]) -> pd.DataFrame:
    if df.empty or "part number" not in df.columns:
        return df
    pn = df["part number"].astype(str).str.lower()
    mask = False
    for m in markers:
        mask = mask | pn.str.contains(m, na=False)
    return df.loc[~mask].reset_index(drop=True)

def drop_description_only_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "description" not in df.columns:
        return df
    other_cols = [c for c in df.columns if c != "description"]
    if not other_cols:
        return df.iloc[0:0].copy()
    desc = df["description"].map(lambda x: "" if pd.isna(x) else str(x).strip())
    others_empty = df[other_cols].apply(
        lambda r: all(("" if pd.isna(v) else str(v).strip()) == "" for v in r.tolist()),
        axis=1
    )
    mask = (desc != "") & (others_empty)
    return df.loc[~mask].reset_index(drop=True)

def pre_merge_cleanup_all_rules(df: pd.DataFrame, config: PipelineConfig) -> pd.DataFrame:
    out = df.copy()

    # remove fully empty rows
    out = out.loc[~out.apply(lambda r: _row_all_empty(r), axis=1)].reset_index(drop=True)

    # remove document metadata / disclaimers
    out = drop_metadata_rows(out)

    if config.drop_pricing_marker_rows_pre_merge:
        out = drop_pricing_marker_rows(out, config.pricing_marker_strings)

    if config.drop_description_only_rows_pre_merge:
        out = drop_description_only_rows(out)

    return out

def normalize_columns_for_merge(df: pd.DataFrame, core_columns: List[str]) -> pd.DataFrame:
    """
    Normalize column names so that dynamic columns align across sections.
    - Core columns -> lowercase canonical
    - Dynamic columns -> cleaned whitespace, original case preserved
    """
    core_lower = {c.lower() for c in core_columns}
    rename_map = {}

    for c in df.columns:
        c_clean = clean_cell(c)
        if c.lower() in core_lower:
            rename_map[c] = c.lower()
        else:
            rename_map[c] = c_clean

    return df.rename(columns=rename_map)


# =============================================================================
# Section processors (kept consistent with your earlier rules)
# =============================================================================
def process_package_section(df: pd.DataFrame, config: PipelineConfig) -> pd.DataFrame:
    if df.empty:
        return df
    out_rows = []
    i, n = 0, len(df)
    parent_required = ["description", "dealer net", "list price", "frt", "installed price"]

    while i < n:
        row = df.iloc[i]
        if _row_all_empty(row):
            i += 1
            continue
        is_parent = all(clean_cell(row.get(c, "")) != "" for c in parent_required)
        if not is_parent:
            i += 1
            continue

        parent = row.copy()
        children_desc = []
        j = i + 1
        while j < n:
            r2 = df.iloc[j]
            if _row_all_empty(r2):
                break
            desc_ok = clean_cell(r2.get("description", "")) != ""
            pn_ok = clean_cell(r2.get("part number", "")) != ""
            nums_empty = all(clean_cell(r2.get(c, "")) == "" for c in ["dealer net", "list price", "frt", "installed price"])
            if desc_ok and pn_ok and nums_empty:
                children_desc.append(clean_cell(r2.get("description", "")))
                if config.keep_package_children_rows:
                    out_rows.append(r2.copy())
            j += 1

        if children_desc:
            includes = "Includes: " + ", ".join([d for d in children_desc if d])
            existing = clean_cell(parent.get("comments", ""))
            parent["comments"] = (existing + " " if existing else "") + includes

        out_rows.append(parent)
        i = j + 1 if j < n else n

    return pd.DataFrame(out_rows).reset_index(drop=True) if out_rows else df.iloc[0:0].copy()

def process_non_package_section(df: pd.DataFrame, config: PipelineConfig) -> pd.DataFrame:
    """
    Process non-package sections.

    Final authoritative rules:
    - Forward-fill ONLY core/pricing columns
    - NEVER forward-fill dynamic applicability columns
    - Exactly ONE row per (description, part number)
    - Prefer rows with applicability signal
    - If no applicability exists, keep first row only
    """

    if df.empty:
        return df

    core_cols = {c.lower() for c in config.required_columns}
    dynamic_cols = [c for c in df.columns if c.lower() not in core_cols]

    rows = []
    prev_kept = None

    for _, row in df.iterrows():
        row = row.copy()

        # ✅ Forward-fill core columns only
        if prev_kept is not None:
            for col in df.columns:
                if col.lower() in core_cols:
                    if clean_cell(row[col]) == "" and clean_cell(prev_kept[col]) != "":
                        row[col] = prev_kept[col]

        keep = True

        if prev_kept is not None:
            same_identity = (
                clean_cell(row.get("description")) == clean_cell(prev_kept.get("description"))
                and clean_cell(row.get("part number")) == clean_cell(prev_kept.get("part number"))
            )

            if same_identity:
                prev_has_app = has_applicability_signal(prev_kept, dynamic_cols)
                curr_has_app = has_applicability_signal(row, dynamic_cols)

                # Drop non-authoritative duplicates
                if not curr_has_app and prev_has_app:
                    keep = False
                elif not curr_has_app and not prev_has_app:
                    keep = False

        if keep:
            rows.append(row)
            prev_kept = row

    return pd.DataFrame(rows).reset_index(drop=True)



# =============================================================================
# Loading data
# =============================================================================


def clean_cell(x: Any) -> str:
    """Clean strings for comparisons; returns '' for nullish."""
    if _is_nullish(x):
        return ""
    s = str(x).replace("\r", " ").replace("\n", " ")
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    return s

def extract_model_metadata_from_df(
    df: pd.DataFrame,
    document_name: str,
    scan_rows: int = 10,
) -> Dict[str, Any]:
    """
    Extract model metadata from a specific worksheet DF.

    Contract:
      - Labels are in column index 2
      - Values are in column index 3 (for model name/year)
      - Publication date:
          * value in LAST ROW, column index 2
          * label in PRECEDING ROW, same column, contains 'Publication Date:'
      - Row positions for model name/year may vary
    """

    model_name = None
    model_year = None
    publication_date = None

    # --------------------------------------------------
    # 1. Model Name & Model Year (scan top block)
    # --------------------------------------------------
    max_rows = min(scan_rows, len(df))

    for r in range(max_rows):
        label = clean_cell(df.iat[r, 2]).lower()
        value = clean_cell(df.iat[r, 3])

        if not label:
            continue

        if label == "model name":
            if not value:
                raise ValueError(f"Model Name value empty at row {r}")
            model_name = value

        elif label == "model year":
            if not value.isdigit():
                raise ValueError(
                    f"Model Year is not numeric at row {r}: '{value}'"
                )
            model_year = int(value)

        if model_name and model_year:
            break

    # --------------------------------------------------
    # 2. Publication Date (strict positional rule)
    # --------------------------------------------------
    last_row_idx = len(df) - 1
    if last_row_idx < 1:
        raise ValueError(
            f"Cannot extract Publication Date from '{document_name}': "
            f"not enough rows."
        )

    pub_label = clean_cell(df.iat[last_row_idx - 1, 2])
    pub_value = clean_cell(df.iat[last_row_idx, 2])

    if "publication date:" not in pub_label.lower():
        raise ValueError(
            f"Expected 'Publication Date:' label at row {last_row_idx - 1}, col 2 "
            f"but found '{pub_label}' in '{document_name}'."
        )

    if not pub_value:
        raise ValueError(
            f"Publication Date value empty at row {last_row_idx}, col 2 "
            f"in '{document_name}'."
        )

    # Strip time portion if present
    # Expected format: YYYY-MM-DD 00:00:00
    publication_date = pub_value.split(" ")[0]

    # --------------------------------------------------
    # 3. Final validation
    # --------------------------------------------------
    missing = []
    if not model_name:
        missing.append("model_name")
    if not model_year:
        missing.append("model_year")
    if not publication_date:
        missing.append("publication_date")

    if missing:
        raise ValueError(
            f"Could not extract metadata {missing} from '{document_name}'. "
            f"Searched rows 0–{max_rows-1} for model fields and last row for publication date."
        )

    return {
        "model_name": model_name,
        "model_year": model_year,
        "publication_date": publication_date,
    }






def load_excel_files_to_dict(directory_path: str) -> Dict[str, Dict[str, pd.DataFrame]]:
    """
    Loads all Excel files (.xlsx, .xlsm) into a dictionary.

    ✅ Loads all worksheets (including hidden)
    ✅ Identifies metadata worksheet via regex
    ✅ Extracts model metadata from that worksheet only
    ✅ Injects authoritative metadata under '__META__'
    """

    if not os.path.exists(directory_path):
        raise ValueError(f"The directory '{directory_path}' does not exist.")
    if not os.path.isdir(directory_path):
        raise ValueError(f"The path '{directory_path}' is not a directory.")

    excel_files = (
        glob(os.path.join(directory_path, "*.xlsx")) +
        glob(os.path.join(directory_path, "*.xlsm"))
    )

    if not excel_files:
        raise FileNotFoundError(f"No Excel files found in '{directory_path}'.")

    result_dict: Dict[str, Dict[str, pd.DataFrame]] = {}

    DATA_FILE_RX = re.compile(r"^\d{2}_?MY_.*_APP_EN$", re.IGNORECASE)

    for file_path in excel_files:
        document_name = os.path.splitext(os.path.basename(file_path))[0]
        model_key = document_name.lower()

        try:
            xls = pd.ExcelFile(file_path, engine="openpyxl")

            workbook: Dict[str, pd.DataFrame] = {}
            metadata_df: Optional[pd.DataFrame] = None

            for sheet_name in xls.sheet_names:
                try:
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    workbook[sheet_name] = df

                    # ✅ Identify metadata worksheet
                    if DATA_FILE_RX.match(sheet_name):
                        metadata_df = df

                except Exception as sheet_err:
                    print(
                        f"⚠️ Skipping sheet '{sheet_name}' in '{document_name}': {sheet_err}"
                    )

            if not workbook:
                print(f"⚠️ No readable sheets in '{document_name}'")
                continue

            if metadata_df is None:
                raise ValueError(
                    f"No worksheet matching metadata pattern found in '{document_name}'"
                )

            # ✅ Extract metadata from the trusted worksheet DF
            model_meta = extract_model_metadata_from_df(
                df=metadata_df,
                document_name=document_name,
            )

            result_dict[model_key] = {
                "__META__": model_meta,
                **workbook,
            }

        except Exception as e:
            print(f"❌ Error reading '{file_path}': {e}")
            continue

    return result_dict


# =============================================================================
# Merge: preserve dynamic columns between core and comments
# =============================================================================
def merge_union_preserve_order(model_key: str, section_dfs: Dict[str, pd.DataFrame], config: PipelineConfig, auditor: ColumnAuditor) -> pd.DataFrame:
    non_empty = {k: v for k, v in section_dfs.items() if v is not None and not v.empty}
    if not non_empty:
        raise DataValidationError("All processed sections are empty; nothing to merge.", model_key=model_key)

    core_lower = [c.lower() for c in config.required_columns]
    core_no_comments = [c for c in core_lower if c != "comments"]

    dynamic_cols: List[str] = []
    seen = set(core_lower)

    for _, df in non_empty.items():
        for c in df.columns:
            cl = str(c).lower()
            if cl not in seen:
                dynamic_cols.append(c)
                seen.add(cl)

    final_cols = core_no_comments + dynamic_cols + (["comments"] if "comments" in core_lower else [])

    aligned = []
    # 1. Normalize column names (non-destructive)
    normalized = []
    for section_key, df in non_empty.items():
        d = normalize_columns_for_merge(df, list(config.required_columns))
        normalized.append(d)

    # 2. Concatenate FIRST (preserves all data)
    merged = pd.concat(normalized, ignore_index=True, sort=False)

    auditor.take(
        model_key,
        "__MERGE__",
        "merge:after_concat_before_reorder",
        merged
    )

    # 3. Build final column order AFTER concat
    core_lower = [c.lower() for c in config.required_columns]
    core_no_comments = [c for c in core_lower if c != "comments"]

    dynamic_cols = []
    seen = set(core_lower)

    for c in merged.columns:
        cl = str(c).lower()
        if cl not in seen:
            dynamic_cols.append(c)
            seen.add(cl)

    final_cols = core_no_comments + dynamic_cols + (["comments"] if "comments" in core_lower else [])

    # 4. Reorder WITHOUT reindexing (no data loss)
    merged = merged.loc[:, final_cols]

    
    # Hard guard: no dynamic column should be entirely NaN
    dynamic_only = [c for c in merged.columns if c.lower() not in core_lower]
    bad = [c for c in dynamic_only if merged[c].notna().sum() == 0]

    if bad:
        raise DataValidationError(
            "Dynamic columns lost all data during merge.",
            model_key=model_key,
            details={"columns": bad}
        )

    auditor.take(
        model_key,
        "__MERGE__",
        "merge:after_column_reorder_no_reindex",
        merged
    )

    auditor.take(model_key, "__MERGE__", "merge:before_pre_merge_cleanup", merged)

    merged = pre_merge_cleanup_all_rules(merged, config)
    auditor.take(model_key, "__MERGE__", "merge:after_pre_merge_cleanup", merged)

    return merged
def drop_leading_columns_if_safe(
    df: pd.DataFrame,
    cols_to_check: List[int],
    core_keywords: List[str],
    auditor: Optional[ColumnAuditor],
    model_key: str
) -> pd.DataFrame:
    """
    Drop leading columns by index ONLY if they do not contain any core keywords.
    """
    out = df.copy()
    keywords_norm = {k.lower() for k in core_keywords}

    drop_cols = []
    for idx in cols_to_check:
        if idx >= out.shape[1]:
            continue

        col = out.iloc[:, idx]
        values = col.astype(str).str.lower()

        contains_core = values.apply(
            lambda v: any(k in v for k in keywords_norm)
        ).any()

        if not contains_core:
            drop_cols.append(out.columns[idx])

    if drop_cols:
        out = out.drop(columns=drop_cols)

    if auditor:
        auditor.take(model_key, "__SHEET__", "sheet:after_drop_leading_columns_0_1", out)

    return out

def drop_metadata_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove rows that contain document-level metadata or disclaimers,
    e.g. 'Publication Date:' or 'All published prices'.
    """
    if df.empty:
        return df

    METADATA_PHRASES = [
        "publication date",
        "all published prices",
        "prices subject to change",
        "pricing subject to change",
        "effective date",
    ]

    def row_is_metadata(row: pd.Series) -> bool:
        for v in row.tolist():
            txt = clean_cell(v).lower()
            if any(p in txt for p in METADATA_PHRASES):
                return True
        return False

    mask = df.apply(row_is_metadata, axis=1)
    return df.loc[~mask].reset_index(drop=True)



# =============================================================================
# Create ready to load file: model specific
# =============================================================================

from collections import defaultdict

def collect_invalid_applicability_by_trim(
    df: pd.DataFrame,
    required_columns: tuple,
) -> dict:
    """
    Count invalid applicability values in dynamic columns,
    grouped by trim and by actual bad value.

    Returns:
    {
        "EX": {"0": 5, "B": 3},
        "LX": {"X": 2}
    }
    """

    required_lower = {c.lower() for c in required_columns}
    metadata_lower = {
        "model_name",
        "model_year",
        "publication_date",
        "make",
    }
    structural_lower = {"trim", "model_number"}

    issues = defaultdict(lambda: defaultdict(int))

    for _, row in df.iterrows():
        trim = row.get("trim")
        if not trim:
            continue

        for col, val in row.items():
            cl = col.lower()

            # ✅ Only dynamic applicability columns
            if (
                cl in required_lower
                or cl in metadata_lower
                or cl in structural_lower
            ):
                continue

            v = clean_cell(val)
            if v and v not in VALID_APPLICABILITY:
                issues[trim][v] += 1

    return {trim: dict(bad_vals) for trim, bad_vals in issues.items()}



def get_trim_columns(df: pd.DataFrame, core_columns: list) -> list:
    core_lower = {c.lower() for c in core_columns}
    metadata_lower = {c.lower() for c in METADATA_COLUMNS}

    return [
        c for c in df.columns
        if c.lower() not in core_lower
        and c.lower() not in metadata_lower
    ]


def row_applies_to_trim(value: str) -> bool:
    return clean_cell(value) in VALID_APPLICABILITY

def build_trim_level_templates(
    merged_df: pd.DataFrame,
    core_columns: list,
) -> Dict[str, pd.DataFrame]:

    trim_cols = get_trim_columns(merged_df, core_columns)
    trim_dfs = {}

    for trim in trim_cols:
        # ✅ KEEP rows with ANY non-empty value
        mask = merged_df[trim].map(lambda v: clean_cell(v) != "")
        df_trim = merged_df.loc[mask].copy()

        if df_trim.empty:
            continue

        # Drop other trim columns
        drop_cols = [c for c in trim_cols if c != trim]
        df_trim = df_trim.drop(columns=drop_cols)

        df_trim["trim"] = trim
        trim_dfs[trim] = df_trim.reset_index(drop=True)

    return trim_dfs



def get_model_number(make: str, model: str, trim: str) -> str:
    # TEMP stub
    return

# def attach_model_number(df: pd.DataFrame) -> pd.DataFrame:
#     df = df.copy()
#     df["model_number"] = df.apply(
#         lambda r: get_model_number(
#             r.get("make"),          # ✅ safe access
#             r.get("model_name"),
#             r.get("trim")
#         ),
#         axis=1
#     )
#     return df

# Added for test, once the model number query system is setup we'll add it 
#  
def attach_model_number(df: pd.DataFrame) -> pd.DataFrame:
    """
    TEMP STUB:
    model_number = trim name
    """
    df = df.copy()
    df["model_number"] = df["trim"]
    return df


def build_model_ready_to_load_df(
    merged_df: pd.DataFrame,
    core_columns: list,
) -> pd.DataFrame:

    trim_dfs = build_trim_level_templates(merged_df, core_columns)


    all_dfs = []
    for trim, df in trim_dfs.items():
        df = attach_model_number(df)
        all_dfs.append(df)

    if not all_dfs:
        return pd.DataFrame()

    return pd.concat(all_dfs, ignore_index=True)

def evaluate_model_status(df: pd.DataFrame) -> Tuple[str, dict]:
    issues = {}

    missing_model_number = df["model_number"].isna().any()
    if missing_model_number:
        issues["missing_model_number"] = True

    required_cols = [
        "description", "dealer net", "list price", "frt", "installed price"
    ]
    missing_required = df[required_cols].isna().any().any()
    if missing_required:
        issues["missing_required_data"] = True

    if not issues:
        return "GREEN", issues
    elif "missing_required_data" in issues:
        return "RED", issues
    else:
        return "BLUE", issues




def save_models_to_excel(
    final_model_dfs: Dict[str, pd.DataFrame],
    report_rows: List[Dict[str, Any]],
    output_dir: str,
):
    
    
    from datetime import datetime

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"Honda_Accessories-Upload_file-{timestamp}.xlsx"

    """
    Save ready-to-load model data into a single Excel file.

    - One worksheet per model
    - One REPORT worksheet
    - Worksheet tab coloring (GREEN / BLUE / RED)
    - Row-level highlighting for data issues
    """

    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, filename)

    # --------------------------------------------------
    # 1. Write raw data with pandas
    # --------------------------------------------------
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for model_key, df in final_model_dfs.items():
            sheet_name = model_key[:31]  # Excel limit
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        report_df = pd.DataFrame(report_rows)
        report_df.to_excel(writer, sheet_name="REPORT", index=False)

    


    # --------------------------------------------------
    # 2. Open workbook for styling
    # --------------------------------------------------
    wb = load_workbook(output_path)

    # Colors
    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    BLUE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    RED = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

    REQUIRED_DATA_COLUMNS = [
        "description",
        "dealer net",
        "list price",
        "frt",
        "installed price",
    ]

    # --------------------------------------------------
    # 3. Style model worksheets
    # --------------------------------------------------
    for report in report_rows:
        model_key = report["model_name"]
        status = report["status"]

        sheet_name = model_key[:31]
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # ---- Tab color ----
        if status == "GREEN":
            ws.sheet_properties.tabColor = "00C6EFCE"
            row_fill = GREEN
        elif status == "BLUE":
            ws.sheet_properties.tabColor = "00BDD7EE"
            row_fill = BLUE
        else:
            ws.sheet_properties.tabColor = "00F4CCCC"
            row_fill = RED

        # ---- Map column names to indexes ----
        headers = [cell.value for cell in ws[1]]
        col_index = {h: i + 1 for i, h in enumerate(headers)}

        # ---- Row-level highlighting ----
        for r in range(2, ws.max_row + 1):
            row_has_issue = False

            for col in REQUIRED_DATA_COLUMNS:
                if col in col_index:
                    val = ws.cell(row=r, column=col_index[col]).value
                    if val is None or str(val).strip() == "":
                        row_has_issue = True
                        break

            if row_has_issue:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = row_fill

    # --------------------------------------------------
    # 4. Style REPORT worksheet
    # --------------------------------------------------
    if "REPORT" in wb.sheetnames:
        ws = wb["REPORT"]

        headers = [cell.value for cell in ws[1]]
        status_col = headers.index("status") + 1

        for r in range(2, ws.max_row + 1):
            status = ws.cell(row=r, column=status_col).value
            fill = GREEN if status == "GREEN" else BLUE if status == "BLUE" else RED

            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = fill

    wb.save(output_path)
    

def create_ready_to_load_data(
    extracted_results: Dict[str, ProcessedRawFileResult],
    config: PipelineConfig,
    output_dir: str,
):
    """
    Phase 3 of the pipeline.

    Responsibilities:
    - Build trim-level templates per model
    - Attach model numbers (stub for now)
    - Merge trims into a ready-to-load model dataset
    - Run final model-level QA
    - Produce Excel output with REPORT worksheet
    """

    COLUMN_NAME_MAPPING_FOR_RIMPORTER = {
        "model_year": "Year",
        "part number": "Part",
        "description": "Description",
        "list price": "Price",
        "frt": "Hours",
        "comments": "Comments",
        "model_number": "Model",
    }

    essential_columns_for_loading = [
        "Year", "Model", "Part", "Description", "Price", "Hours", "Comments"
    ]

    final_model_dfs: Dict[str, pd.DataFrame] = {}
    report_rows: List[Dict[str, Any]] = []

    for model_key, result in extracted_results.items():
        merged_df = result.merged_df

        # --------------------------------------------------
        # Defaults (must exist for all code paths)
        # --------------------------------------------------
        final_df = pd.DataFrame()
        filtered_df_for_load = pd.DataFrame()
        status = "RED"
        issues = ""
        records_by_trim = ""
        invalid_applicability_by_trim = {}
        integrity_issues_parts = []

        # --------------------------------------------------
        # Safety: empty model
        # --------------------------------------------------
        if merged_df.empty:
            report_rows.append({
                "model_name": model_key,
                "model_year": None,
                "publication_date": None,
                "trims_processed": 0,
                "trim_names": "",
                "records_by_trim": "",
                "rows_total": 0,
                "missing_model_number": True,
                "data_integrity_issues": "no_data",
                "status": "RED",
            })
            final_model_dfs[model_key] = merged_df
            continue

        # --------------------------------------------------
        # 1. Build trim-level templates
        # --------------------------------------------------
        trim_dfs = build_trim_level_templates(
            merged_df=merged_df,
            core_columns=list(config.required_columns),
        )

        # ✅ ASSERT: invalid values must survive to QA phase
        for trim, df_trim in trim_dfs.items():
            bad = df_trim[trim].map(
                lambda v: clean_cell(v) not in ("", *VALID_APPLICABILITY)
            )
            if bad.any():
                print(f"[QA DEBUG] Invalid values preserved for trim '{trim}'")

        # Metadata survival assert
        for trim, df_trim in trim_dfs.items():
            for col in ["model_name", "model_year", "publication_date", "make"]:
                assert col in df_trim.columns, (
                    f"{model_key}: metadata column '{col}' was dropped in trim '{trim}'"
                )

        # Trim isolation asserts
        for trim, df_trim in trim_dfs.items():
            assert df_trim["trim"].nunique() == 1
            assert df_trim["trim"].iloc[0] == trim

        # --------------------------------------------------
        # 2. Attach model number (stub = trim)
        # --------------------------------------------------
        model_trim_dfs = []
        for _, df_trim in trim_dfs.items():
            df_trim = attach_model_number(df_trim)
            model_trim_dfs.append(df_trim)

        if not model_trim_dfs:
            issues = "no_trims_after_filter"
            status = "RED"
        else:
            # --------------------------------------------------
            # 3. Merge all trims
            # --------------------------------------------------
            final_df = pd.concat(model_trim_dfs, ignore_index=True)

            # --------------------------------------------------
            # 4. Final QA evaluation
            # --------------------------------------------------
            status, issues_dict = evaluate_model_status(final_df)
            
            

            # --------------------------------------------------
            # Data quality: invalid applicability values (by trim)
            # --------------------------------------------------
            
            if not final_df.empty:
                invalid_applicability_by_trim = collect_invalid_applicability_by_trim(
                    df=final_df,
                    required_columns=config.required_columns,
                )

            
            if invalid_applicability_by_trim:
                print("Invalid applicability detected:", invalid_applicability_by_trim)

            if invalid_applicability_by_trim:
                status = "RED"
            
            integrity_issues_parts = []

            if invalid_applicability_by_trim:
                for trim, bad_vals in invalid_applicability_by_trim.items():
                    inner = ", ".join(f"{k}:{v}" for k, v in bad_vals.items())
                    integrity_issues_parts.append(f"{trim}={{ {inner} }}")


            # Combine issues
            all_issues = []

            if issues_dict:
                all_issues.extend(
                    f"{k}={v}" if isinstance(v, int) else k
                    for k, v in issues_dict.items()
                )

            if integrity_issues_parts:
                all_issues.append(
                    "invalid_applicability: " + "; ".join(integrity_issues_parts)
                )

            issues = " | ".join(all_issues)

            # --------------------------------------------------
            # Records per trim
            # --------------------------------------------------
            records_by_trim_map = (
                final_df.groupby("trim").size().sort_index().to_dict()
            )
            records_by_trim = ", ".join(
                f"{trim}={count}" for trim, count in records_by_trim_map.items()
            )

            # --------------------------------------------------
            # Prepare data for load
            # --------------------------------------------------
            df_to_upload = final_df.copy()
            df_to_upload.rename(
                columns=COLUMN_NAME_MAPPING_FOR_RIMPORTER,
                inplace=True,
            )

            for col in essential_columns_for_loading:
                assert col in df_to_upload.columns, (
                    f"{model_key}: missing column '{col}' for load"
                )

            filtered_df_for_load = df_to_upload[essential_columns_for_loading]

        # --------------------------------------------------
        # Save model DF
        # --------------------------------------------------
        final_model_dfs[model_key] = filtered_df_for_load

        # --------------------------------------------------
        # Build REPORT row
        # --------------------------------------------------
        trim_names = (
            sorted(final_df["trim"].unique().tolist())
            if not final_df.empty else []
        )

        publication_date = (
            final_df["publication_date"].iloc[0]
            if not final_df.empty else None
        )

        report_rows.append({
            "model_name": final_df["model_name"].iloc[0]
                if not final_df.empty else model_key,
            "model_year": final_df["model_year"].iloc[0]
                if not final_df.empty else None,
            "publication_date": publication_date,
            "trims_processed": len(trim_names),
            "trim_names": ", ".join(trim_names),
            "records_by_trim": records_by_trim,
            "rows_total": len(final_df),
            "missing_model_number": (
                final_df["model_number"].isna().any()
                if not filtered_df_for_load.empty else True
            ),
            "data_integrity_issues": issues,
            "status": status,
        })

    # ------------------------------------------------------
    # Final safety assert
    # ------------------------------------------------------
    for model_key, df in final_model_dfs.items():
        if not df.empty:
            for col in essential_columns_for_loading:
                assert col in df.columns

    # ------------------------------------------------------
    # Save Excel output + REPORT worksheet
    # ------------------------------------------------------
    save_models_to_excel(
        final_model_dfs=final_model_dfs,
        report_rows=report_rows,
        output_dir=output_dir,
    )

    


# =============================================================================
# Main: process one model
# =============================================================================

def extract_model_file(
    model_key: str,
    excel_dict: Dict[str, Dict[str, pd.DataFrame]],
    make: str,
    config: PipelineConfig,
)-> ProcessedRawFileResult:

    auditor = make_auditor(config)

    if model_key not in excel_dict:
        raise DataValidationError("Model key not found in excel_dict.", model_key=model_key)

    model_dict = excel_dict[model_key]


    en_sheet = find_english_worksheet_name(model_dict, config, model_key)


    if en_sheet is None:
        raise WorksheetNotFoundError(
            f"English sheet not found for model '{model_key}' using pattern {config.english_pattern_template}",
            model_key=model_key,
            details={"available": list(model_dict.keys())}
        )

    english_df = model_dict[en_sheet]



    if not isinstance(english_df, pd.DataFrame):
        raise DataValidationError("English sheet is not a DataFrame.", model_key=model_key)

    # ✅ Drop leading columns 0 and 1 if safe
    core_keywords = list(config.required_columns) + [config.description_header_text]

    
    english_df = drop_leading_columns_if_safe(
        english_df,
        cols_to_check=[0, 1],
        core_keywords=core_keywords,
        auditor=auditor,
        model_key=model_key
    )

    if config.strip_strings_on_load:
        english_df = strip_strings_in_df(english_df)

    auditor.take(model_key, "__SHEET__", "sheet:english_loaded", english_df)

    sections = extract_sections(model_key, english_df, config, auditor)

    processed_sections: Dict[str, pd.DataFrame] = {}
    ready: Dict[str, pd.DataFrame] = {}

    for section_key, sec_df in sections.items():

        # ✅ Header fix (header guard lives INSIDE this function)
        fixed = fix_section_header_new_spec(
            model_key,
            section_key,
            sec_df,
            config,
            auditor
        )

        # ✅ Section-specific processing
        if section_key == "1.0 Packages and Kits":
            processed = process_package_section(fixed, config)
        else:
            processed = process_non_package_section(fixed, config)

        # ✅ GUARD: section processing must not wipe dynamic data
        guard_dynamic_columns(
            before_df=fixed,
            after_df=processed,
            core_columns=list(config.required_columns),
            model_key=model_key,
            section_key=section_key,
            stage="after_section_processing",
        )

        # Enforce required pricing fields
        validate_required_columns_not_empty(
            processed,
            required_cols=[
                "description",
                "dealer net",
                "list price",
                "frt",
                "installed price",
            ],
            model_key=model_key,
            section_key=section_key,
        )

        # Enforce refined uniqueness
        validate_pricing_identity_uniqueness(
            processed,
            model_key=model_key,
            section_key=section_key,
            core_columns=list(config.required_columns),
        )
        
        auditor.take(model_key, section_key, "section:after_processing", processed)

        # ✅ Pre-merge cleanup
        cleaned = pre_merge_cleanup_all_rules(processed, config)

        # ✅ GUARD: cleanup must not wipe dynamic data
        guard_dynamic_columns(
            before_df=processed,
            after_df=cleaned,
            core_columns=list(config.required_columns),
            model_key=model_key,
            section_key=section_key,
            stage="after_pre_merge_cleanup",
        )
        

        auditor.take(model_key, section_key, "section:after_pre_merge_cleanup", cleaned)

        processed_sections[section_key] = cleaned
        ready[section_key] = cleaned


    # ✅ Merge (merge-level guards already handled inside merge function)
    merged_df = merge_union_preserve_order(
        model_key,
        ready,
        config,
        auditor
    )
    # model_name, model_year, publication_date    
    merged_df["model_name"] = excel_dict[model_key]["__META__"]["model_name"]
    merged_df["model_year"] = excel_dict[model_key]["__META__"]["model_year"]
    merged_df["publication_date"] = excel_dict[model_key]["__META__"]["publication_date"]
    merged_df["make"] = make

    
    
    # ✅ HARD ASSERTS – extraction contract
    required_meta_cols = ["model_name", "model_year", "publication_date", "make"]

    for col in required_meta_cols:
        assert col in merged_df.columns, (
            f"extract_model_file contract violated: missing required metadata column '{col}' "
            f"in merged_df")


    return ProcessedRawFileResult(
        model_key=model_key,
        merged_df=merged_df,
        english_sheet_name=en_sheet,
        section_dfs=processed_sections
    )


def data_extract_orchestrator(
    excel_dict: Dict[str, Dict[str, pd.DataFrame]],
    make: str,
    config: PipelineConfig,
):
    """
    Executes the accessory pipeline for all models in excel_dict.

    Returns:
        results: Dict[str, ProcessedRawFileResult]
        failures: Dict[str, Exception]
    """
    results = {}
    failures = {}

    for model_key in sorted(excel_dict.keys()):
        print("\n" + "=" * 120)
        print(f"🚀 Processing model: {model_key}")
        print("=" * 120)

        try:
            
            res = extract_model_file(
                model_key=model_key,
                excel_dict=excel_dict,
                make=make,
                config=config,
            )

            results[model_key] = res
            

            print(f"✅ SUCCESS: {model_key}")
            print(f"   Rows: {len(res.merged_df)}")
            print(f"   Columns: {len(res.merged_df.columns)}")

        except Exception as e:
            failures[model_key] = e
            print(f"❌ FAILURE: {model_key}")
            print(str(e))


    print("\n" + "=" * 120)
    print("✅ RUN COMPLETE")
    print(f"   Successful models: {len(results)}")
    print(f"   Failed models: {len(failures)}")
    print("=" * 120)

    return results, failures


def run_pipeline(
    input_dir: str,
    output_dir: str,
    make: str,
    config: PipelineConfig,
):

    """
    Top-level pipeline orchestrator.

    Responsibilities:
    - Load Excel files
    - Extract model data (Phase 1)
    - Build trim-level ready-to-load datasets (Phase 2)
    - Save Excel output + REPORT (Phase 3)
    """

    print("\n" + "=" * 120)
    print("📥 Loading Excel files")
    print("=" * 120)

    excel_dict = load_excel_files_to_dict(input_dir)

    print("\n" + "=" * 120)
    print("🔎 Extracting model data")
    print("=" * 120)

    extracted_results, failures = data_extract_orchestrator(
        excel_dict=excel_dict,
        make=make,
        config=config,
    )

    

    if failures:
        print("\n⚠️ Some models failed extraction:")
        for model_key, err in failures.items():
            print(f"  - {model_key}: {err}")

    if not extracted_results:
        raise RuntimeError("No models were successfully extracted. Pipeline aborted.")

    print("\n" + "=" * 120)
    print("🧩 Building trim-level lists and exporting Excel")
    print("=" * 120)

    create_ready_to_load_data(
        extracted_results=extracted_results,
        config=config,
        output_dir=output_dir,
    )

    print("\n✅ Pipeline completed successfully")


# %%
RAW_DATA_DIR = pl.Path("../landing_zone/")
READY_TO_UPLOAD_FILES_DIR = pl.Path("../ready_to_upload/")

HONDA_DATA_PATH = RAW_DATA_DIR / "2026" / "Honda" / "2026-05-04"
HONDA_READY_TO_UPLOAD_DIR = READY_TO_UPLOAD_FILES_DIR / "2026" / "Honda"
 

# %%

# %%

excel_dict = load_excel_files_to_dict(HONDA_DATA_PATH)


# %%

# # Test data_extract_orchestrator()
# results, _ = data_extract_orchestrator(excel_dict, make="Honda", config=config)

# for model_key, res in results.items():
#     df = res.merged_df
#     assert "model_name" in df.columns
#     assert "model_year" in df.columns
#     assert "publication_date" in df.columns
#     assert "make" in df.columns


# %%

# excel_dict["accord"]["26MY_ACCORD_APP_FR"].iloc[9:]
 

# %%
# config = PipelineConfig(
#     strict_mode=True,
#     audit_enabled=False,
# )


# # results, _ = data_extract_orchestrator(excel_dict, make="Honda", config=config)

# # for model_key, res in results.items():
# #     df = res.merged_df
# #     assert "model_name" in df.columns
# #     assert "model_year" in df.columns
# #     assert "publication_date" in df.columns
# #     assert "make" in df.columns

# run_pipeline(
#     input_dir=HONDA_DATA_PATH,
#     output_dir=HONDA_READY_TO_UPLOAD_DIR,
#     make="Honda",
#     config=config,
# )


# %%

# %% [markdown]
# ### Adding the French version

# %%
# french 

# %%
primary_cols_to_check = {
    
}


# %%

def load_excel_sheets(excel_path: str) -> Dict[str, pd.DataFrame]:
    """
    Load all sheets from an Excel file into a dictionary.

    Args:
        excel_path (str): Path to the Excel file.

    Returns:
        Dict[str, pd.DataFrame]: Dictionary mapping sheet name to DataFrame.
    """
    # sheet_name=None tells pandas to load all sheets
    sheets = pd.read_excel(excel_path, sheet_name=None)
    return sheets
 
def drop_rows_empty_in_columns(
    df: pd.DataFrame,
    columns: List[str]
) -> pd.DataFrame:
    """
    Drop rows where ALL specified columns are empty (NaN / None).

    Args:
        df: Input DataFrame
        columns: Columns to check for emptiness

    Returns:
        Filtered DataFrame
    """
    return df.dropna(subset=columns, how="all")


def deduplicate_by_completeness(
    df: pd.DataFrame,
    key_cols: List[str],
    required_cols: List[str],
) -> pd.DataFrame:
    """
    Deduplicate a DataFrame by keeping, per key, the row that has data
    in all required_cols (or the most complete one).

    If ambiguity remains (tie), raises ValueError.

    Args:
        df: Input DataFrame (e.g. df_a)
        key_cols: Columns that define uniqueness (lookup keys)
        required_cols: Columns that must have data (columns to populate)

    Returns:
        Deduplicated DataFrame
    """

    # Count non-null values in required columns per row
    df = df.copy()
    df["_completeness_score"] = df[required_cols].notna().sum(axis=1)

    # Sort so best row per key comes first
    df_sorted = df.sort_values(
        by=key_cols + ["_completeness_score"],
        ascending=[True] * len(key_cols) + [False]
    )

    # Keep the best row per key
    deduped = df_sorted.drop_duplicates(subset=key_cols, keep="first")

    # ---- Ambiguity check ----
    # If multiple rows per key have the same max completeness → ambiguous
    max_scores = (
        df.groupby(key_cols)["_completeness_score"]
        .max()
        .reset_index(name="max_score")
    )

    merged = df.merge(max_scores, on=key_cols, how="left")

    ambiguous = (
        merged["_completeness_score"] == merged["max_score"]
    ) & merged.duplicated(subset=key_cols, keep=False)

    if ambiguous.any():
        amb_keys = (
            merged.loc[ambiguous, key_cols]
            .drop_duplicates()
            .sort_values(key_cols)
        )

        raise ValueError(
            "Ambiguous duplicate keys found after completeness filtering.\n"
            f"Keys:\n{amb_keys.to_string(index=False)}"
        )

    # Cleanup
    deduped = deduped.drop(columns="_completeness_score")

    return deduped




# %%

# loaded the processed excel file containing ready to load data

ready_to_load_excelFiles_dict = load_excel_sheets(HONDA_READY_TO_UPLOAD_DIR / 'Honda_Accessories-Upload_file-2026-05-04_12-17-58.xlsx')


# %%

ready_to_load_excelFiles_dict["accord"].columns


# %%
# 
my_counter =0 
for sheet in ready_to_load_excelFiles_dict:
    if "Model" not in ready_to_load_excelFiles_dict[sheet].columns:
        print(f"WARNING: 'Model' column not found in sheet '{sheet}'")
        continue  
    number_of_unique_trims = ready_to_load_excelFiles_dict[sheet]["Model"].nunique()  
    my_counter += number_of_unique_trims * 2


# %%
my_counter


def time_saved(number_of_trims: int) -> float:
    # Assuming 30 minutes saved per trim and 2 trims per model
    minutes_saved_per_model = number_of_trims * 45
    hours_saved = minutes_saved_per_model / 60
    return hours_saved

time_saved(my_counter)


# %%
# # Get all unique partNumbers in the ready to load df file

# to_load_accord_df = ready_to_load_excelFiles_dict["accord"]

# uniq_partNumbers = to_load_accord_df["Part"]

# french_df_accord = excel_dict["accord"]["26MY_ACCORD_APP_FR"]

# %%


def drop_duplicates_keep_complete(
    df: pd.DataFrame,
    part_col: List[str],
    required_cols: List[str],
    model: str,
) -> pd.DataFrame:
    """
    Drop duplicate rows based on part_col, keeping the row that has
    data in ALL required_cols.

    Fails if:
      - No row per key has all required columns populated
      - More than one row per key has all required columns populated

    Args:
        df: Input DataFrame
        part_col: Columns defining the key
        required_cols: Columns that must all be non-null
        model: Model name (for logging / context)

    Returns:
        Deduplicated DataFrame
    """

    df = df.copy()

    # Normalize empties
    df[required_cols] = df[required_cols].replace("", pd.NA)

    dupe_mask = df.duplicated(subset=part_col, keep=False)

    if not dupe_mask.any():
        return df

    # Log affected keys
    dup_keys = (
        df.loc[dupe_mask, part_col]
        .drop_duplicates()
        .sort_values(part_col)
    )

    print(
        f"[INFO] Duplicates found for model={model}\n"
        f"Keys affected:\n{dup_keys.to_string(index=False)}"
    )

    kept_rows = []
    errors = []

    for _, group in df.groupby(part_col, dropna=False):
        # Rows where ALL required columns are present
        complete_rows = group[group[required_cols].notna().all(axis=1)]

        if len(complete_rows) == 1:
            kept_rows.append(complete_rows.iloc[0])
        elif len(complete_rows) == 0:
            errors.append(
                f"No complete row found for key:\n{group[part_col].iloc[0].to_dict()}"
            )
        else:
            errors.append(
                f"Multiple complete rows found for key:\n{group[part_col].iloc[0].to_dict()}"
            )

    if errors:
        raise ValueError(
            f"[{model}] Duplicate resolution failed:\n" + "\n".join(errors)
        )

    result = pd.DataFrame(kept_rows).reset_index(drop=True)
    return result


def assert_no_missing_values(df, columns, model):
    # Normalize empty strings to NA
    check_df = df[columns].replace("", pd.NA)

    # Rows where ANY required column is missing
    mask = check_df.isna().any(axis=1)

    if mask.any():
        missing_rows = df.loc[mask, columns]

        print(f"\n❌ [{model}] Rows with missing values in required columns:")
        print(missing_rows.to_string(index=True))
        print(f"missing part numbers : {missing_rows["Part"].tolist()}")
        raise ValueError(
            f"[{model}] Missing values detected in columns: {columns}"
        )
    

def assert_merge_integrity(
    df: pd.DataFrame,
    column_pairs: Dict[str, str],
    model: str,
) -> None:
    """
    Assert that merged columns were populated correctly.

    For each (left_col, right_col) pair:
        - If left_col has a value, right_col must also have a value.

    Args:
        df: DataFrame after merge
        column_pairs: {left_col: right_col}
        model: Context / model name for error messages
    """

    violations = []

    for left_col, right_col in column_pairs.items():
        if left_col not in df.columns or right_col not in df.columns:
            raise KeyError(
                f"[{model}] Columns '{left_col}' or '{right_col}' not found in DataFrame"
            )

        left_vals = df[left_col].replace("", pd.NA)
        right_vals = df[right_col].replace("", pd.NA)

        mask = left_vals.notna() & right_vals.isna()

        if mask.any():
            bad_rows = df.loc[mask, [left_col, right_col]]
            bad_rows["_failed_pair"] = f"{left_col} -> {right_col}"
            violations.append(bad_rows)

    if violations:
        error_df = pd.concat(violations, axis=0)

        print(f"\n❌ [{model}] Merge integrity check failed.")
        print("Rows where left key has data but merged value is missing:\n")
        print(error_df.to_string(index=True))

        raise ValueError(
            f"[{model}] Merge lost data for one or more key column pairs"
        )
    


def assert_column_pair_presence_match(
    df: pd.DataFrame,
    column_pairs: Dict[str, str],
    model: str,
) -> None:
    """
    Assert that for each column pair (col_a, col_b):
    - Both have values OR
    - Both are missing

    If only one has a value, print rows for auditing and raise an error.
    """

    violations = []

    for col_a, col_b in column_pairs.items():
        if col_a not in df.columns or col_b not in df.columns:
            raise KeyError(
                f"[{model}] Columns '{col_a}' or '{col_b}' not found in DataFrame"
            )

        a_vals = df[col_a].replace("", pd.NA)
        b_vals = df[col_b].replace("", pd.NA)

        # XOR → True when exactly one side has data
        mask = a_vals.notna() ^ b_vals.notna()

        if mask.any():
            bad_rows = df.loc[mask, [col_a, col_b]].copy()
            bad_rows["failed_pair"] = f"{col_a} ↔ {col_b}"
            violations.append(bad_rows)

    if violations:
        audit_df = pd.concat(violations, axis=0)

        print(f"\n❌ [{model}] Column presence mismatch detected")
        print("Rule: each column pair must either BOTH have values or BOTH be empty\n")
        print(audit_df.to_string(index=True))

        raise ValueError(
            f"[{model}] Column presence mismatch found for one or more column pairs"
        )
    

def get_single_complete_row_df(
    df: pd.DataFrame,
    required_cols: List[str],
) -> pd.DataFrame:
    """
    Return a DataFrame containing exactly ONE row that has data
    in ALL required_cols.

    Raises:
        ValueError if zero or more than one qualifying row exists.
    """

    if df.empty:
        raise ValueError("DataFrame is empty")

    # Normalize empty strings to NA
    check_df = df[required_cols].replace("", pd.NA)

    # Rows where ALL required columns are populated
    mask = check_df.notna().all(axis=1)
    matches = df.loc[mask]

    if len(matches) == 1:
        return matches.reset_index(drop=True)

    if len(matches) == 0:
        raise ValueError(
            f"No row found with all required columns populated: {required_cols}"
        )

    # len(matches) > 1
   
    if len(matches) > 1:

        # print("❌ Multiple complete rows found")
        # print("Required columns:", required_cols)
        # print("Candidate rows:")

        selected_row = matches.iloc[:1]
        return selected_row.reset_index(drop=True)

        
        # display(matches)

        # raise ValueError("Ambiguous complete rows")



def filter_for_french_file(df):

    all_columns = list(df.columns)

    cols_to_remove = ["Description", "Comments"]

    for col in cols_to_remove:
        if col in all_columns:
            all_columns.remove(col)

    df = df[all_columns]
    

    cols_rename_map = {"Description_fr":"Description", "Comments_fr":"Comments"}

    filtered_FR_cols_final = df.rename(columns=cols_rename_map)
    
    return filtered_FR_cols_final


# %%


def replace_nan_with_empty(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """
    Replace NaN values and the literal string "nan" (case-insensitive)
    with empty strings in specified columns.

    Only replaces when the cell value is exactly "nan".

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame
    columns : list[str]
        Columns to apply the replacement to

    Returns
    -------
    pd.DataFrame
        Transformed DataFrame
    """
    df = df.copy()

    for col in columns:
        df[col] = (
            df[col]
            # .replace(to_replace=r"^(?i)nan$", value="", regex=True)
            .replace(to_replace=r"(?i)^nan$", value="", regex=True)
            .fillna("")
        )

    return df

# %%


def populate_french_data(french_df, target_df, model):

    # Param : Part Number, French dataframe
    # Return : a sub df with french data to be populated on to the ready to load df
    
    rename_french_key_cols = {  
                                "Unnamed: 3":"Part",
                                "Unnamed: 5":"Price",
                                "Unnamed: 6":"Hours",
                                "Unnamed: 2":"Description_fr",
                                "Unnamed: 16":"Comments_fr",    
                            }
    
    french_df = french_df.rename(columns=rename_french_key_cols)


    match_cols = {
                    "Part":"Part",
                    "Price":"Price",
                    "Hours":"Hours",
                    }
    
    source_cols = ["Description_fr", "Comments_fr"]

    french_df_cols = list(match_cols.keys())
    uniq_parNumbers = target_df["Part"]
    original_eng_columns = list(target_df.columns)
    # Filter the list to matching records
    filtered_df = french_df[
                        (french_df["Part"].isin(uniq_parNumbers))
                        & french_df[french_df_cols].notna().all(axis=1)
                            ]    
    
    # filtered_df = drop_duplicates_keep_one(filtered_df, "Part", model)
    
        # assert no missing values
    # assert_no_missing_values(filtered_df, list(match_cols.keys()), model)

    part_col = "Part"
    required_cols_for_french_df = list(match_cols.keys()) + ["Description_fr"] 
    # # merge matching columns
    
    # filtered_df = drop_duplicates_keep_complete(filtered_df, part_col, required_cols_for_french_df, model)

    # assert no missing values
    # assert_no_missing_values(filtered_df, list(match_cols.keys()), model)
    
    a_keys = list(match_cols.keys())
    b_keys = list(match_cols.values())
     
    # assert filtered_df["Part"].is_unique, (
    #     f"Duplicate values found in filtered_df['Part']:\n"
    #     f"{filtered_df.loc[filtered_df['Part'].duplicated(), 'col'].to_string(index=False)}"
    #     )
    
    # ---------- 2. Prepare lookup table ----------
    lookup = (
        filtered_df[a_keys + source_cols]
    )
    added_french_cols_df = target_df.copy()
    # ---------- 3. Merge with explicit left_on / right_on ----------
    processed_parts = set()
    records_missing_in_french_version = pd.DataFrame()

    for idx, row in added_french_cols_df.iterrows():
        part = row["Part"]

        # Skip if this Part was already handled
        if part in processed_parts:
            continue

        processed_parts.add(part)

        french_row = french_df[
            (french_df["Part"] == row["Part"])
            # & (french_df["Price"] == row["Price"])
        ]

        
        if french_row.empty:
            continue

        if isinstance(french_row, type(None)):
            display(row)
            raise ValueError(f"Duplicate French rows found for Part={part}")
        
        if len(french_row) > 1:
            french_row = get_single_complete_row_df(french_row, required_cols_for_french_df)
            
        # if len(french_row) > 1:
        #     display(french_row)
        #     raise ValueError(f"Duplicate French rows found for Part={part}")

        desc_fr = str(french_row.iloc[0]["Description_fr"])
        comm_fr = str(french_row.iloc[0]["Comments_fr"])

        # ✅ Populate ALL rows with the same Part at once
        mask = added_french_cols_df["Part"] == part

        added_french_cols_df.loc[mask, "Description_fr"] = desc_fr
        added_french_cols_df.loc[mask, "Comments_fr"] = comm_fr
        
        added_french_cols_df =  replace_nan_with_empty(added_french_cols_df, ["Description_fr", "Comments_fr"])
        
    
    # Get search result
    result = added_french_cols_df[
        added_french_cols_df["Description_fr"].isna()
    ].copy()   # <-- important

    # Add model column
    result["model_file"] = model  # e.g. "Honda_FR"

    # Append (via concat)
    records_missing_in_french_version = pd.concat([records_missing_in_french_version, result], ignore_index=True)

    # filter for French columns only
    
    if not isinstance(added_french_cols_df, pd.DataFrame):
        raise ValueError(f"Final populated value not a pandas DataFrame {model}")

    french_output_df = filter_for_french_file(added_french_cols_df)
    
    return {"df_with_FR_cols":{model:french_output_df}, "record_missing_in_FR_file": records_missing_in_french_version}
    # return added_french_cols_df, recs_missing_in_french_version

 
# merged_df_en_and_fr = populate_french_data(french_df_accord, to_load_accord_df)

# %%
ready_to_load_excelFiles_dict.keys()

# %%
french_file_RX = re.compile(r"^\d{2}_?MY_.*_APP_FR$", re.IGNORECASE)
excel_dict.keys()

# print(french_file_RX.match("26MY_CIV4D_APP_FR"))

# %%



# %%
# Select legitimate models and remove REPORT for search purposes
all_models = list(ready_to_load_excelFiles_dict.keys())
all_models.remove("REPORT")
french_version_df_dict = {}
missing_in_french_files = pd.DataFrame()

# loop through all models to get the french data sheet form the original data feed
for model in all_models:
   french_file_RX = re.compile(r"^\d{2}_?MY_.*_APP_FR$", re.IGNORECASE)
   if model=='civic4d':
       model = 'civic 4d'

   if model=='cvic5d':
      model = 'civic 5d'
       
   for sheet_name in excel_dict[model].keys():
       if french_file_RX.match(sheet_name):
         french_df  = excel_dict[model][sheet_name]
         if model== 'civic 4d':
               model ='civic4d'
         if model=='civic 5d':
            model = 'cvic5d'
         english_df = ready_to_load_excelFiles_dict[model]
         populated_results = populate_french_data(french_df, english_df, sheet_name)

         french_version_df_dict[model] = populated_results["df_with_FR_cols"][sheet_name]
         missing_in_french_files = pd.concat([missing_in_french_files, populated_results["record_missing_in_FR_file"]], ignore_index=True)

         

# %%

# %%
# ready_to_load_excelFiles_dict["accord"]["Comments"].str.isalnum()

# %%
excel_dict['prologue'].keys()


# %%
french_version_df_dict.keys()

# %%
# excel_dict["pilot"]["26MY_Pilot_APP_FR"]

# %%
models_list = ['civic 5d', 'accord', 'civic 4d', 'crv', 'hrv', 'odyssey', 'passport', 'pilot', 'prelude', 'prologue', 'ridgeline']

idx = 9
model = models_list[idx]

model_sheet = models_list[idx]
# model_sheet = 'hr-v'

FR_dic_model_sheet = models_list[idx]
# print(model_sheet)

sheet_FR = f'26MY_{model_sheet.capitalize()}_APP_FR'
# sheet_FR = f'26MY_{model_sheet.upper()}_APP_FR'
# sheet_FR = f'26MY_{model.upper()}_APP_FR'
# sheet_FR = f'26MY_{model}_APP_FR'



french_df_accord = excel_dict[model][sheet_FR]

filtered_df = french_version_df_dict[FR_dic_model_sheet][
    (french_version_df_dict[FR_dic_model_sheet]["Description"].isna()) 
    | (french_version_df_dict[FR_dic_model_sheet]["Description"]=='')
    | (french_version_df_dict[FR_dic_model_sheet]["Description"]=='0') # Data discrepencies whereby description in the FR version is set to zero, which may imply missing data or part unavailable for french
    | (french_version_df_dict[FR_dic_model_sheet]["Part"]=='')
    | (french_version_df_dict[FR_dic_model_sheet]["Price"]=='')
    | (french_version_df_dict[FR_dic_model_sheet]["Hours"]=='')
    | (french_version_df_dict[FR_dic_model_sheet]["Model"]=='')    
]
print("Records missing data")
print(f'Model : {model}')
display(filtered_df)                              

all_price_items = set(filtered_df["Price"]) 
print("Source records records")
columns_to_print = ["Unnamed: 1","Unnamed: 2","Unnamed: 3", "Unnamed: 5", "Unnamed: 6", "Unnamed: 16"]
french_df_accord[french_df_accord["Unnamed: 5"].isin(all_price_items)][columns_to_print]


# %%
french_df_accord[french_df_accord["Unnamed: 5"].str.contains('305', case=False, na=False)]

# %%

# %%
# len(english_df)

# %%
# # populated_df[populated_df["Description"].str.contains("Package|kit|set", case=False)]
# populated_df[
#     (populated_df["Description_fr"].isnull()) 
#     & ~(populated_df["Description"].isnull())          
#              ]




# %%
# # populated_df[populated_df["Description_fr"].str.contains("", case=False)]
# populated_df[populated_df["Part"]=="TUCH-B575P-A1"]



# %%
# def search_part(part_number):

#     display(french_df[french_df["Unnamed: 3"]==part_number])

    
# search_part("TUCH-B575P-A1")

# %%

# def save_dfs_to_excel(
#     sheets: Dict[str, pd.DataFrame],
#     output_dir: str,
#     engine: str = "openpyxl",
# ) -> None:
#     """
#     Save multiple DataFrames to a single Excel file,
#     using dict keys as worksheet names.

#     Args:
#         sheets: Dictionary of {sheet_name: DataFrame}
#         output_path: Full path to the output Excel file
#         engine: Excel writer engine (default: openpyxl)

#     Returns:
#         None
#     """


#     from datetime import datetime

#     timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
#     filename = f"FR_Honda_Accessories-Upload_file-{timestamp}.xlsx"

#     os.makedirs(output_dir, exist_ok=True)
#     file_output_path = os.path.join(output_dir, filename)

#     if not sheets:
#         raise ValueError("sheets dictionary is empty")
    
#     with pd.ExcelWriter(file_output_path, engine=engine) as writer:
#         for sheet_name, df in sheets.items():
#             df.to_excel(
#                 writer,
#                 sheet_name=str(sheet_name),
#                 index=False
#             )



# %%

# %%


def save_dfs_to_excel(
    sheets: Dict[str, pd.DataFrame],
    output_dir: str,
    engine: str = "openpyxl",
) -> None:
    """
    Save multiple DataFrames to a single Excel file.

    Expected input:
        sheets = {
            "SheetName1": df1,
            "SheetName2": df2,
            ...
        }
    """
    from datetime import datetime

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"FR_Honda_Accessories-Upload_file-{timestamp}.xlsx"

    # ---------- Validation ----------
    if not isinstance(sheets, dict):
        raise TypeError(
            f"sheets must be a dict of {{sheet_name: DataFrame}}, got {type(sheets).__name__}"
        )

    if not sheets:
        raise ValueError("sheets dict is empty")

    parsed_sheets = []

    for sheet_name, df in sheets.items():
        if not isinstance(sheet_name, str):
            raise TypeError(f"Sheet name '{sheet_name}' is not a string")

        if not isinstance(df, pd.DataFrame):
            raise TypeError(f"Value for sheet '{sheet_name}' is not a DataFrame")

        parsed_sheets.append((sheet_name, df))

    # ---------- File prep ----------
    os.makedirs(output_dir, exist_ok=True)
    file_output_path = os.path.join(output_dir, filename)

    # ---------- Write ----------
    with pd.ExcelWriter(file_output_path, engine=engine) as writer:
        wrote_any = False

        for sheet_name, df in parsed_sheets:
            if df.empty:
                continue  # optional: skip empty DataFrames

            safe_sheet_name = sheet_name[:31]  # Excel sheet name limit

            df.to_excel(
                writer,
                sheet_name=safe_sheet_name,
                index=False
            )
            wrote_any = True

        if not wrote_any:
            raise ValueError(
                "No non-empty DataFrames were written; Excel requires at least one visible sheet"
            )




# def save_dfs_to_excel(
#     sheets: List[Dict[str, pd.DataFrame]],
#     output_dir: str,

#     engine: str = "openpyxl",
# ) -> None:
#     """
#     Save multiple DataFrames to a single Excel file.

#     Expected input:
#         sheets = [
#             {"SheetName1": df1},
#             {"SheetName2": df2},
#             ...
#         ]
#     """
#     from datetime import datetime

#     timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
#     filename = f"FR_Honda_Accessories-Upload_file-{timestamp}.xlsx"
    
    
#     # ---------- Validation ----------
#     if not isinstance(sheets, list):
#         raise TypeError(
#             f"sheets must be a list of dicts, got {type(sheets).__name__}"
#         )

#     if not sheets:
#         raise ValueError("sheets list is empty")

#     parsed_sheets = []

#     for i, item in enumerate(sheets):
#         if not isinstance(item, dict):
#             raise TypeError(f"Item at index {i} is not a dict")

#         if len(item) != 1:
#             raise ValueError(
#                 f"Item at index {i} must contain exactly one {{sheet_name: DataFrame}} pair"
#             )

#         sheet_name, df = next(iter(item.items()))

#         if not isinstance(sheet_name, str):
#             raise TypeError(f"Sheet name at index {i} is not a string")

#         if not isinstance(df, pd.DataFrame):
#             raise TypeError(f"Value for sheet '{sheet_name}' is not a DataFrame")

#         parsed_sheets.append((sheet_name, df))

#     # ---------- File prep ----------
#     os.makedirs(output_dir, exist_ok=True)
#     file_output_path = os.path.join(output_dir, filename)

#     # ---------- Write ----------
#     with pd.ExcelWriter(file_output_path, engine=engine) as writer:
#         wrote_any = False

#         for sheet_name, df in parsed_sheets:
#             if df.empty:
#                 continue  # optional: skip empty DataFrames

#             safe_sheet_name = sheet_name[:31]  # Excel sheet name limit

#             df.to_excel(
#                 writer,
#                 sheet_name=safe_sheet_name,
#                 index=False
#             )
#             wrote_any = True

#         if not wrote_any:
#             raise ValueError(
#                 "No non-empty DataFrames were written; Excel requires at least one visible sheet"
#             )




# %%
# # Select legitimate models and remove REPORT for search purposes
# all_models = list(ready_to_load_excelFiles_dict.keys())
# all_models.remove("REPORT")
# french_version_df_dict = []

# # loop through all models to get the french data sheet form the original data feed
# for model in all_models:
#     french_file_RX = re.compile(r"^\d{2}_?MY_.*_APP_FR$", re.IGNORECASE)
#     for sheet_name in excel_dict[model].keys():
#        if french_file_RX.match(sheet_name):
#          french_df  = excel_dict[model][sheet_name]
#          english_df = ready_to_load_excelFiles_dict[model]
        
#          french_version_df_dict.append()
         
            
        
#     # excel_dict[model].keys()
#     # french_df = excel_dict[model][""]

# %%
# french_version_df_dict[1]["26MY_CIV4D_APP_FR"]

# %%

# non_null_columns = list(final_fr_df.columns) 

# non_null_columns.remove("Comments")
# # check if any of the non-null columns have any null value 
# # assert_no_missing_values(final_fr_df, non_null_columns, model)

# %%

# %%
# save_dfs_to_excel(french_version_df_dict, output_dir = HONDA_READY_TO_UPLOAD_DIR)

# %%
# Accord,Civic Sedan,Civic Hatch,CR-V,HR-V,Odyssey,Passport,Pilot,Prelude,Prologue,Ridgeline

# %%
